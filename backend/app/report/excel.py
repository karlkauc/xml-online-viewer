"""Render a validation result as a styled .xlsx workbook (openpyxl)."""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import __version__
from app.parser.validate import ValidationResponse

# Application metadata shown in the report and stored in the workbook's
# document properties.
APP_NAME = "FundsXML Online Validator"
APP_AUTHOR = "Karl Kauc"
APP_URL = "https://xml-viewer.status20.net"

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_META_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_MONO_FONT = Font(name="Consolas")
# Inline font for the offending line inside the context cell (bold red).
_ERROR_LINE_INLINE = InlineFont(rFont="Consolas", b=True, color="B91C1C")

_COLUMNS = [
    ("#", 6),
    ("Severity", 12),
    ("Zeile", 8),
    ("Spalte", 8),
    ("XML-Pfad", 50),
    ("Nachricht", 90),
    ("Typ", 28),
    ("Domain", 18),
    ("Kontext (Zeile davor / Fehlerzeile / danach)", 70),
]
# 1-based column index of the context cell.
_CONTEXT_COL = 9


def _context_cell(prev: str | None, err: str | None, nxt: str | None) -> CellRichText | None:
    """Build a single rich-text cell holding the line before, the error line
    (bold red) and the line after, separated by newlines."""
    if err is None:
        return None
    segments: list[object] = []
    if prev is not None:
        segments.append(prev + "\n")
    segments.append(TextBlock(_ERROR_LINE_INLINE, err))
    if nxt is not None:
        segments.append("\n" + nxt)
    return CellRichText(*segments)


def _context_lines(lines: list[str], line: int | None) -> tuple[str | None, str | None, str | None]:
    """Return (line above, error line, line below) for a 1-based line number,
    or (None, None, None) if unavailable."""
    if not line or line < 1 or line > len(lines):
        return (None, None, None)
    idx = line - 1
    prev = lines[idx - 1] if idx - 1 >= 0 else None
    nxt = lines[idx + 1] if idx + 1 < len(lines) else None
    return (prev, lines[idx], nxt)


def build_report(
    result: ValidationResponse,
    *,
    xml_filename: str,
    xsd_filename: str,
    reformatted_xml: str = "",
) -> bytes:
    """Return the bytes of an .xlsx workbook describing ``result``.

    ``reformatted_xml`` is the pretty-printed document the error line numbers
    index into; it is used to emit the line before, the error line (highlighted)
    and the line after each error, to ease locating the problem.
    """
    lines = reformatted_xml.splitlines()
    wb = Workbook()
    ws = wb.active
    ws.title = "Validierung"

    generated_at = datetime.now(timezone.utc)

    # Workbook document properties (visible under File ▸ Info / Properties).
    wb.properties.title = f"{APP_NAME} – Validierungsreport"
    wb.properties.creator = APP_NAME
    wb.properties.lastModifiedBy = APP_NAME
    wb.properties.subject = f"Validierung {xml_filename} gegen {xsd_filename}"
    wb.properties.description = (
        f"Erzeugt von {APP_NAME} {__version__} ({APP_URL}); Autor {APP_AUTHOR}."
    )
    wb.properties.keywords = "FundsXML, XSD, Validierung"
    wb.properties.created = generated_at
    wb.properties.modified = generated_at

    ws["A1"] = f"{APP_NAME} – Validierungsreport"
    ws["A1"].font = _TITLE_FONT

    meta = [
        ("Applikation", APP_NAME),
        ("Version", __version__),
        ("Autor", APP_AUTHOR),
        ("URL", APP_URL),
        ("", ""),
        ("XML-Datei", xml_filename),
        ("XSD-Schema", xsd_filename),
        ("Erstellt", generated_at.strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Ergebnis", "GÜLTIG" if result.is_valid else "UNGÜLTIG"),
        ("Anzahl Fehler", len(result.errors)),
    ]
    row = 3
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = _META_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1

    header_row = row + 1
    for col_idx, (title, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for i, err in enumerate(result.errors, start=1):
        r = header_row + i
        prev_line, err_line, next_line = _context_lines(lines, err.line)
        values = [
            i,
            err.severity,
            err.line,
            err.column,
            err.path,
            err.message,
            err.type_name,
            err.domain,
            _context_cell(prev_line, err_line, next_line),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            if col_idx in (5, 6):  # path, message wrap
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            elif col_idx == _CONTEXT_COL:  # context: monospace, wrapped lines
                cell.font = _MONO_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    last_row = header_row + len(result.errors)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(_COLUMNS))}{max(last_row, header_row)}"
    )
    # Widen the first column so the metadata labels are readable.
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 16)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
